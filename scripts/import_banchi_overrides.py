import json
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_INPUTS = [
    Path("未プロット中古マンション一覧_updated.xlsx"),
    Path("未プロット中古マンション一覧.xlsx"),
]
OUTPUT = Path("used_banchi_overrides.json")


def clean(value):
    return str(value or "").strip()


def main():
    candidates = [path for path in DEFAULT_INPUTS if path.exists()]
    if not candidates:
        raise FileNotFoundError("未プロット中古マンション一覧.xlsx が見つかりません")

    best_source = None
    best_overrides = {}

    for source in candidates:
        wb = load_workbook(source, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        columns = {header: index + 1 for index, header in enumerate(headers)}

        id_col = columns.get("物件番号")
        banchi_cols = [
            col for col in [columns.get("確認後の番地"), columns.get("正規化番地")]
            if col
        ]
        if not id_col or not banchi_cols:
            continue

        overrides = {}
        for row_index in range(2, ws.max_row + 1):
            property_id = clean(ws.cell(row_index, id_col).value)
            banchi = ""
            for banchi_col in banchi_cols:
                banchi = clean(ws.cell(row_index, banchi_col).value)
                if banchi:
                    break
            if property_id and banchi:
                overrides[property_id] = banchi

        if len(overrides) >= len(best_overrides):
            best_source = source
            best_overrides = overrides

    if best_source is None:
        raise ValueError("Excelに「物件番号」または番地列がありません")

    OUTPUT.write_text(json.dumps(best_overrides, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"wrote {OUTPUT} entries={len(best_overrides)} source={best_source}")


if __name__ == "__main__":
    main()
