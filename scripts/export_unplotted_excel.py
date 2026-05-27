import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SOURCE = Path("used_properties.js")
OUTPUT = Path("未プロット中古マンション一覧.xlsx")


def load_used_properties(path):
    text = path.read_text(encoding="utf-8")
    prefix = "const usedProperties = "
    start = text.index(prefix) + len(prefix)
    end = text.rindex(";")
    return json.loads(text[start:end])


def main():
    rows = [
        item for item in load_used_properties(SOURCE)
        if item.get("lat") is None or item.get("lng") is None
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "未プロット一覧"

    headers = [
        "物件番号",
        "物件名",
        "行政区",
        "住所（町丁目まで）",
        "元番地",
        "正規化番地",
        "間取り",
        "価格",
        "専有面積",
        "最寄駅",
        "徒歩",
        "築年月",
        "CSVファイル",
        "未プロット理由",
        "確認後の番地",
        "確認メモ",
    ]
    ws.append(headers)

    for item in rows:
        ws.append([
            item.get("id", ""),
            item.get("name", ""),
            item.get("ward", ""),
            item.get("address", ""),
            item.get("banchiOriginal", ""),
            item.get("banchi", ""),
            item.get("layout", ""),
            item.get("priceLabel", ""),
            item.get("areaLabel", ""),
            item.get("station", ""),
            item.get("walk", ""),
            item.get("built", ""),
            item.get("sourceFile", ""),
            item.get("geocodeStatus", ""),
            "",
            "",
        ])

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = [14, 34, 10, 32, 14, 14, 12, 12, 12, 14, 10, 12, 16, 18, 18, 30]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    output = OUTPUT
    try:
        wb.save(output)
    except PermissionError:
        output = OUTPUT.with_name(f"{OUTPUT.stem}_updated{OUTPUT.suffix}")
        wb.save(output)
    print(f"wrote {output} rows={len(rows)}")


if __name__ == "__main__":
    main()
